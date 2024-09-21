import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
import win32com.client as win32
import os

# Load the existing workbook and select the specified worksheet
input_excel_file = "test.xlsx"  # Ensure this file is present in the same directory
wb = openpyxl.load_workbook(input_excel_file)
ws = wb["sample"]  # Replace 'Data' with your sheet name

# Create a BarChart for 'Passed'
bar_chart = BarChart()
bar_chart.title = "CLEC DATA"
# bar_chart.x_axis.title = "Date"
# bar_chart.y_axis.title = "Passed"
# bar_chart.y_axis.majorGridlines = None

# Define data for the BarChart
categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
values = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
bar_chart.add_data(values, titles_from_data=True)
bar_chart.set_categories(categories)

# Create a LineChart for 'Blocked'
line_chart = LineChart()
line_chart.y_axis.title = "Blocked"
line_chart.y_axis.axId = 200
line_chart.y_axis.crosses = "max"
line_values = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
line_chart.add_data(line_values, titles_from_data=True)
line_chart.set_categories(categories)

# Combine the BarChart and LineChart
bar_chart.y_axis.crosses = "autoZero"
bar_chart += line_chart

# Add the combined chart to the worksheet
ws.add_chart(bar_chart, "E5")

# Save the workbook with the chart included
temp_excel_file = "clec_data_with_chart.xlsx"
wb.save(temp_excel_file)

# Open Excel application
excel = win32.Dispatch('Excel.Application')
excel.Visible = False

# Open the Excel file
wb = excel.Workbooks.Open(os.path.abspath(temp_excel_file))
ws = wb.Worksheets("sample")

# Access the chart object
chart = ws.ChartObjects(1)  # Access the first chart object

# Save the chart as a PNG image
image_file = os.path.abspath("clec_data.png")
chart.Chart.Export(image_file)

# Close the workbook and quit Excel
wb.Close(False)
excel.Quit()

# Optionally, delete the temporary Excel file if no longer needed
# os.remove(temp_excel_file)

# Send the image via email using Outlook
outlook = win32.Dispatch('outlook.application')

# Create a new mail item
mail = outlook.CreateItem(0)

# Set the email properties
mail.Subject = "CLEC Data Chart"
mail.Body = "Hi Team,\n\nPlease find the attached CLEC data chart image.\n\nBest regards,\n[Your Name]"
mail.To = "pamedi@ibasis.net"  # Replace with your group's email
mail.CC = "msaleem@ibasis.net"     # Optional: Replace with any CC email addresses
# mail.BCC = "bcc@example.com"   # Optional: Replace with any BCC email addresses

# Attach the image file
mail.Attachments.Add(image_file)

# Send the email
mail.Send()

print("Email sent successfully.")
